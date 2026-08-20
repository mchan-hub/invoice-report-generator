import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter
import re
from io import BytesIO

st.set_page_config(page_title="Invoice Management Report Generator", page_icon="🧾", layout="wide")

st.title("🧾 Invoice Management Report Generator")
st.markdown("Upload your **Report Template** and at least one of the data files (**Xero** or **NetSuite**) to generate the consolidated report.")

@st.cache_data
def load_data(file):
    if file.name.lower().endswith('.csv'):
        return pd.read_csv(file)
    else:
        try:
            return pd.read_excel(file)
        except Exception:
            file.seek(0)
            try:
                return pd.read_csv(file, sep='\t')
            except:
                file.seek(0)
                return pd.read_html(file)[0]

def clean_amount(val):
    if pd.isna(val): return 0.0
    val_str = str(val)
    cleaned = re.sub(r'[^\d.-]', '', val_str)
    try:
        if cleaned.count('.') > 1:
            parts = cleaned.rsplit('.', 1)
            cleaned = parts[0].replace('.', '') + '.' + parts[1]
        return float(cleaned) if cleaned else 0.0
    except ValueError:
        return 0.0

col1, col2, col3 = st.columns(3)
with col1: xero_file = st.file_uploader("1. Upload Xero Data (Optional)", type=["csv", "xls", "xlsx"])
with col2: ns_file = st.file_uploader("2. Upload NetSuite Data (Optional)", type=["csv", "xls", "xlsx"])
with col3: template_file = st.file_uploader("3. Upload Report Template (Required)", type=["xlsx"])

st.markdown("---")
st.subheader("💱 匯率設定 (FX Rates to USD)")
st.write("請輸入以下貨幣兌換成 **USD** 嘅匯率 (例如 1 HKD = 0.128 USD)。數值支援至 6 個小數位：")

target_currencies = ['HKD', 'SGD', 'MYR', 'IDR', 'BRL', 'GBP', 'AED', 'EUR', 'USD']
fx_rates = {}

fx_cols = st.columns(3)
for i, cur in enumerate(target_currencies):
    with fx_cols[i % 3]:
        default_val = 1.000000 if cur == 'USD' else 0.000000
        fx_rates[cur] = st.number_input(f"{cur} to USD Rate", value=default_val, step=0.000100, format="%.6f")

st.write("---")
st.write("請輸入計算最底總結算 (Total HKD) 時所用嘅匯率：")
usd_to_hkd = st.number_input("USD to HKD Exchange Rate", value=7.80, step=0.01)
st.markdown("---")

if st.button("🚀 Generate Report", type="primary", use_container_width=True):
    if template_file and (xero_file or ns_file):
        try:
            with st.spinner("Processing data & mapping columns..."):
                dfs_to_combine = []
                
                if xero_file:
                    xero_df = load_data(xero_file)
                    xero_mapped = pd.DataFrame({
                        'Location': 'Xero',
                        'Invoice Number': xero_df.get('InvoiceNumber', ''),
                        'Client Name': xero_df.get('ContactName', ''),
                        'Currency': xero_df.get('Currency', ''),
                        'Amount': xero_df.get('Total', 0),
                        'Due date': pd.to_datetime(xero_df.get('DueDate', ''), format='mixed', dayfirst=True, errors='coerce'),
                        'Company id': xero_df.get('Reference', '')
                    })
                    dfs_to_combine.append(xero_mapped)
                
                if ns_file:
                    ns_df = load_data(ns_file)
                    ns_mapped = pd.DataFrame({
                        'Location': ns_df.get('Subsidiary', 'NetSuite'),
                        'Invoice Number': ns_df.get('Document Number', ''),
                        'Client Name': ns_df.get('Name', ''),
                        'Currency': ns_df.get('Currency', ''),
                        'Amount': ns_df.get('Amount (Foreign Currency)', '').apply(clean_amount),
                        'Due date': pd.to_datetime(ns_df.get('Due Date/Receive By', ''), format='mixed', errors='coerce'),
                        'Contact Owner': ns_df.get('Sales Rep', ''),
                        'Company id': ns_df.get('External ID', '')
                    })
                    dfs_to_combine.append(ns_mapped)
                
                combined_df = pd.concat(dfs_to_combine, ignore_index=True)
                combined_df['Due date'] = combined_df['Due date'].dt.strftime('%Y-%m-%d').fillna('')
                combined_df['fx rate'] = combined_df['Currency'].map(fx_rates).fillna(1.0)

                wb = openpyxl.load_workbook(template_file)
                ws = wb.active

                header_row = 2
                headers = {str(ws.cell(row=header_row, column=c).value).strip(): c for c in range(1, ws.max_column + 1) if ws.cell(row=header_row, column=c).value}
                
                col_map = {
                    'Location': headers.get('Location'),
                    'Invoice Number': headers.get('Invoice Number'),
                    'Client Name': headers.get('Client Name'),
                    'Currency': headers.get('Currency'),
                    'Amount': headers.get('Amount'),
                    'Due date': headers.get('Due date'),
                    'Contact Owner': headers.get('Contact Owner'),
                    'Company id': headers.get('Company id'),
                    'fx rate': headers.get('fx rate')
                }
                
                ref_row = 3
                formula_cols = [c for c in range(1, ws.max_column + 1) if ws.cell(row=ref_row, column=c).data_type == 'f']

                start_row = 3
                for r_idx, row in combined_df.iterrows():
                    current_row = start_row + r_idx
                    
                    for col_name, col_idx in col_map.items():
                        if col_idx:
                            ws.cell(row=current_row, column=col_idx).value = row[col_name]
                    
                    if current_row > ref_row:
                        for c_idx in formula_cols:
                            ref_cell = ws.cell(row=ref_row, column=c_idx)
                            new_cell = ws.cell(row=current_row, column=c_idx)
                            new_cell.value = Translator(ref_cell.value, origin=ref_cell.coordinate).translate_formula(new_cell.coordinate)
                
                last_row = start_row + len(combined_df) - 1
                summary_row_usd = last_row + 2
                summary_row_hkd = last_row + 3
                
                usd_without_int_col = headers.get('USD Without interest')
                usd_with_int_col = headers.get('USD With interest')
                
                if usd_without_int_col and usd_with_int_col:
                    ws.cell(row=summary_row_usd, column=usd_without_int_col - 1).value = "Total (USD):"
                    ws.cell(row=summary_row_usd, column=usd_without_int_col).value = f"=SUM({get_column_letter(usd_without_int_col)}3:{get_column_letter(usd_without_int_col)}{last_row})"
                    ws.cell(row=summary_row_usd, column=usd_with_int_col).value = f"=SUM({get_column_letter(usd_with_int_col)}3:{get_column_letter(usd_with_int_col)}{last_row})"
                    
                    ws.cell(row=summary_row_hkd, column=usd_without_int_col - 1).value = "Total (HKD):"
                    ws.cell(row=summary_row_hkd, column=usd_without_int_col).value = f"={get_column_letter(usd_without_int_col)}{summary_row_usd} * {usd_to_hkd}"
                    ws.cell(row=summary_row_hkd, column=usd_with_int_col).value = f"={get_column_letter(usd_with_int_col)}{summary_row_usd} * {usd_to_hkd}"

                    # 強制設定呢兩欄為「千位分隔符，並保留兩位小數」嘅純數字格式，唔顯示貨幣符號
                    number_format = '#,##0.00'
                    for r in range(start_row, summary_row_hkd + 1):
                        ws.cell(row=r, column=usd_without_int_col).number_format = number_format
                        ws.cell(row=r, column=usd_with_int_col).number_format = number_format

                output = BytesIO()
                wb.save(output)
                output.seek(0)
                
                st.success("✅ Report generated successfully!")
                st.download_button(
                    label="📥 Download Updated Invoice Report",
                    data=output,
                    file_name="Updated_Invoice_Management_Report.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        except Exception as e:
            st.error(f"❌ 發生錯誤，請檢查檔案格式: {e}")
    else:
        st.warning("⚠️ 請上傳「Report Template」以及最少一份數據檔案 (Xero 或 NetSuite)。")
